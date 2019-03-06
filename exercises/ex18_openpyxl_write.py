from openpyxl import *
wb = Workbook()
ws = wb.active

# arp_table as dictionary of dictionaries
#

arp_table = {'0080:2101:cd07': {'ip': '10.10.197.8', 'age': '22'},
             '005e:97b2:0854': {'ip': '10.10.197.4', 'age': '119'},
             '0081:2778:a9b9': {'ip': '10.10.197.7', 'age': '37'}
            }

row = 1
for mac, dict in arp_table.items():
    ws.cell(row=row, column=1).value = mac
    ws.cell(row=row, column=2).value = dict['ip']
    ws.cell(row=row, column=3).value = dict['age']
    row += 1

wb.save('arp_table.xlsx')
